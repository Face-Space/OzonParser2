import logging
import openpyxl
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, output_dir: Path, filename: str):
       self.output_dir = output_dir
       self.filename = filename
       self.filepath = output_dir / f"{filename}.xlsx"


    def export_results(self, data: dict, selected_fields: list = None) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ozon Products"

            # Маппинг полей
            field_mapping = {
                'article': ('Артикул', lambda p: p.get('article', '')),
                'name': ('Название товара', lambda p: p.get('name', '')),
                'seller_name': ('Продавец', lambda p: p.get('seller', {}).get('name', '')),
                'company_name': ('Название компании', lambda p: p.get('seller', {}).get('company_name', '')),
                'inn': ('ИНН', lambda p: p.get('seller', {}).get('inn', '')),
                'card_price': ('Цена карты', lambda p: p.get('card_price', 0)),
                'price': ('Цена', lambda p: p.get('price', 0)),
                'original_price': ('Старая цена', lambda p: p.get('original_price', 0)),
                'product_url': ('Ссылка товара', lambda p: p.get('product_url', '')),
                'image_url': ('Изображение', lambda p: p.get('image_url', '')),
                'orders_count': ('Заказов', lambda p: p.get('seller', {}).get('orders_count', '')),
                'reviews_count': ('Отзывов', lambda p: p.get('seller', {}).get('reviews_count', '')),
                'average_rating': ('Рейтинг', lambda p: p.get('seller', {}).get('average_rating', '')),
                'working_time': ('Работает с', lambda p: p.get('seller', {}).get('working_time', ''))
            }

            # Используем выбранные поля или все по умолчанию
            if selected_fields:
                headers = [field_mapping[field][0] for field in selected_fields if field in field_mapping]
                field_extractors = [field_mapping[field][1] for field in selected_fields if field in field_mapping]
                # headers — список названий колонок (первых элементов кортежей из field_mapping) для выбранных полей.
                # field_extractors — список лямбда-функций (вторых элементов кортежей) для выбранных полей.
            else:
                # По умолчанию: название товара, название компании, ссылка на товар и изображение
                default_fields = ['name', 'company_name', 'product_url', 'image_url']
                headers = [field_mapping[field][0] for field in default_fields]
                field_extractors = [field_mapping[field][1] for field in default_fields]

            # Стили
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Данные
            for row, product in enumerate(data.get('products', []), 2):
                row_data = [extractor(product) for extractor in field_extractors]
                # В цикле берётся каждый product из списка словарей data['products'].
                # Для каждого поля из field_extractors (а это список лямбда-функций из field_mapping)
                # вызывается функция с аргументом product

                # Цветовая индикация успешности
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') if product.get('success') else PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    if not product.get('success'):
                        cell.fill = fill

            # Ширина колонок (адаптивная)
            default_widths = {'Артикул': 12, 'Название товара': 40, 'Продавец': 25, 'Название компании': 30,
                              'ИНН': 15,
                              'Цена карты': 12, 'Цена': 12, 'Старая цена': 12, 'Ссылка товара': 50,
                              'Изображение': 50, 'Заказов': 12, 'Отзывов': 12, 'Рейтинг': 12, 'Работает с': 15}

            for col, header in enumerate(headers, 1):
                width = default_widths.get(header, 15)
                ws.column_dimensions[get_column_letter(col)].width = width
                # get_column_letter(col) — функция из openpyxl, которая преобразует номер столбца (число) в букву
                # столбца Excel (например, 1 → 'A', 2 → 'B' и т.д.).
                # У свойства ws.column_dimensions для столбца с буквой, соответствующей col, устанавливается атрибут
                # .width равный найденной ширине.
                # Это задаёт ширину каждого столбца на листе Excel.

            # Высота строк
            for row in range(1, len(data.get('products', [])) + 2):
                ws.row_dimensions[row].height = 20





